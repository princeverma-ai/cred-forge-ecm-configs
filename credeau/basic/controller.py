import os
from openpyxl import load_workbook

def get_filtered_names(file_name, sheet_name, name_column, filter_column, filter_value):
    # Get the absolute path of the Excel file
    script_dir = os.path.dirname(os.path.abspath(__file__))  # Directory of the Python script
    file_path = os.path.join(script_dir, file_name)  # Full path to the Excel file

    # Debugging: print the file path to verify it's correct
    # print(f"Checking file at path: {file_path}")

    # Verify if the file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found at path: {file_path}")

    # Load the workbook and select the sheet
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Initialize an empty list to store the filtered names
    names = []

    # Loop through rows in the sheet, starting from the second row (to skip headers)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[name_column - 1]  # Adjust for zero-based indexing
        filter_value_in_row = row[filter_column - 1]  # Adjust for zero-based indexing
        
        # print(f"Row: {row}")  # Debugging to see row contents
        
        # Check if the filter condition is met (e.g., if the carrierNameOrCollection value matches the filter_value)
        if filter_value_in_row == filter_value:
            # print(f"Name: {name}, Carrier Value: {filter_value_in_row}")
            names.append(name)  # Add the name to the list

    return names

def name(names):

    for index, name_item in enumerate(names, start=1):  # `enumerate` will give the index and the name
        print(f"{index} {name_item}")
    

# Example usage
file_name = r"A:\task\convert name to list\object.xlsx"
sheet_name = "Sheet1"       # Replace with your sheet name
name_column = 6             # Column number for the 'name' column (adjusted based on your data)
filter_column = 7          # Column number for 'carrierNameOrCollection'
filter_value = 1            # Filter condition value (e.g., 1 for matching carrierNameOrCollection)

try:
    filtered_names = get_filtered_names(file_name, sheet_name, name_column, filter_column, filter_value)
    # print("Filtered names:", filtered_names)
    name(filtered_names)
except FileNotFoundError as e:
    print(e)
except Exception as e:
    print("An error occurred:", e)
